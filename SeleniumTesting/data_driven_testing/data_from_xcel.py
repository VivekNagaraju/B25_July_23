import openpyxl

path=r"C:\Users\admin\Desktop\vivek.xlsx"
my_workbook=openpyxl.load_workbook(path)
active_sheet=my_workbook.active

rows=active_sheet.max_row
columns=active_sheet.max_column

print(rows)
print(columns)

for r in range(2, rows+1):
    username=active_sheet.cell(row=r, column=1).value
    password=active_sheet.cell(row=r, column=2).value
    print("username:",username)
    print("password:",password)